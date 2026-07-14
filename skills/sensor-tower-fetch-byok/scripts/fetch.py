"""
Sensor Tower API 一键取数 + 投资口径调整（BYOK 自带 Key 版，自包含）。

与团队版功能完全相同，唯一区别：**不内置任何 token**。每位使用者需自行提供自己的
Sensor Tower API key。给「产品 + 国家 + 日期区间」，直接调 ST API 拉下载量 / 收入（可选 DAU），
套可选口径（query.json 的 caliber: net=ST原始净收入不÷0.7 / gross=÷0.7还原毛支出；CN上修默认关），重算 RPD/ARPDAU，出多 sheet Excel。不依赖其它 skill。口径见 数据口径说明.md。

鉴权（按优先级，三选一）：
  1) 环境变量 ST_AUTH_TOKEN
  2) 同 skill 根目录 config.local.json 里的 auth_token（已 gitignore，不会被提交，推荐放自己的 key）
  3) 同 skill 根目录 config.json 里的 auth_token（默认是占位符，本版本不内置真实 token）

用法：
    python3 fetch.py <query.json> --resolve-only      # 先确认产品匹配
    python3 fetch.py <query.json> [<output.xlsx>]      # 正式拉数 + 调整 + 出 Excel

query.json 字段见 SKILL.md。products/start_date/end_date 必填，其余取 config 默认。
"""

from __future__ import annotations

import json
import os
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import requests
from openpyxl.utils import get_column_letter

# ============== 口径常量 ==============
# 收入口径由 query.json 的 "caliber" 字段控制(默认 net):
#   "net"  = ST 原始净收入(已扣平台分成+税费),不 ÷0.7 —— 公司战略会分析默认用此
#   "gross"= ÷0.7 还原成玩家毛支出 —— 用于与 gross 口径外部数据对齐
# CN 区上修默认关闭(CN_*_MULT=1.0); 如需开启改下方常量。详见 数据口径说明.md
CN_CODES = {"CN"}
GROSS_SHARE = 0.7         # gross 口径还原系数(平台 30% 分成)
CN_REV_MULT = 1.0         # CN 经验上修(默认关=1.0; 开启=2.2)
CN_DL_MULT = 1.0          # CN 经验上修(默认关=1.0; 开启=4)
CN_DAU_MULT = 1.0         # CN 经验上修(默认关=1.0; 开启=3)
# ===========================================================================

BASE = "https://api.sensortower.com"
SKILL_ROOT = Path(__file__).resolve().parent.parent
CONFIG_PATH = SKILL_ROOT / "config.json"
LOCAL_CONFIG_PATH = SKILL_ROOT / "config.local.json"   # gitignored 本地覆盖（放自己的 key）
WW_CODE = "WW"
OTHER_LABEL = "OTHER"
REQUEST_TIMEOUT = 90
MAX_RETRIES = 4
OS_SALES_FIELDS = {
    "unified": (("unified_units",), ("unified_revenue",)),
    "ios": (("iphone_units", "ipad_units"), ("iphone_revenue", "ipad_revenue")),
    "android": (("android_units",), ("android_revenue",)),
}
OS_DAU_FIELDS = {
    "unified": ("android_users", "ipad_users", "iphone_users"),
    "ios": ("iphone_users", "ipad_users"),
    "android": ("android_users",),
}

TOKEN_HELP = (
    "未找到 Sensor Tower API token。本版本不内置任何 key，请用你自己的 ST API token，三选一：\n"
    "  1) 环境变量：  export ST_AUTH_TOKEN=<你的token>\n"
    "  2) 本地配置：  在本 skill 根目录建 config.local.json（已 gitignore，不会被提交）：\n"
    '                 {\"auth_token\": \"<你的token>\"}\n'
    "  3) 直接改 config.json 的 auth_token（注意：该文件会被提交，别把私钥推到公共仓库）。\n"
    "  · API token 在 Sensor Tower 账号后台的 API/Integrations 设置里获取（需相应数据订阅权限）。"
)


# ----------------------- 配置 / 鉴权 -----------------------

def load_config() -> dict:
    cfg: dict = {}
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            cfg.update(json.load(f))
    if LOCAL_CONFIG_PATH.exists():     # 本地覆盖优先
        with open(LOCAL_CONFIG_PATH, "r", encoding="utf-8") as f:
            cfg.update(json.load(f))
    return cfg


def get_token(cfg: dict) -> str:
    tok = os.environ.get("ST_AUTH_TOKEN") or cfg.get("auth_token")
    if not tok or str(tok).startswith("<"):
        raise SystemExit(TOKEN_HELP)
    return tok


# ----------------------- 带退避重试的 GET -----------------------

def st_get(token: str, path: str, params: dict) -> list | dict:
    url = f"{BASE}{path}"
    p = dict(params)
    p["auth_token"] = token
    last_err: Optional[Exception] = None
    for attempt in range(MAX_RETRIES):
        try:
            r = requests.get(url, params=p, timeout=REQUEST_TIMEOUT)
        except Exception as e:
            last_err = e
        else:
            if r.status_code == 200:
                return r.json()
            if r.status_code in (401, 403):
                raise SystemExit(
                    f"鉴权失败 HTTP {r.status_code}：你的 token 无效或无此数据权限。\n{TOKEN_HELP}")
            if r.status_code == 422:
                raise SystemExit(f"参数错误 HTTP 422：{path}\n   {r.text[:300]}")
            if r.status_code not in (429, 500, 502, 503, 504):
                raise SystemExit(f"HTTP {r.status_code}：{path}\n   {r.text[:300]}")
            last_err = RuntimeError(f"HTTP {r.status_code}")
        if attempt < MAX_RETRIES - 1:
            wait = 2 ** (attempt + 1)
            print(f"   [retry] {path} 失败（{last_err}），{wait}s 后重试 ({attempt + 1}/{MAX_RETRIES - 1})")
            time.sleep(wait)
    raise SystemExit(f"重试 {MAX_RETRIES} 次仍失败：{path}（{last_err}）")


# ----------------------- 产品名解析 -----------------------

def _looks_like_unified_id(s: str) -> bool:
    return len(s) == 24 and all(c in "0123456789abcdef" for c in s.lower())


def resolve_product(token: str, entry) -> dict:
    explicit_id = None
    term = None
    if isinstance(entry, dict):
        explicit_id = entry.get("unified_id") or entry.get("id")
        term = entry.get("name")
    elif isinstance(entry, str):
        if _looks_like_unified_id(entry):
            explicit_id = entry
        else:
            term = entry
    else:
        raise ValueError(f"无法识别的 products 条目：{entry!r}")

    if explicit_id:
        return {"unified_id": explicit_id, "name": term or explicit_id,
                "publisher": "(显式 id)", "release_date": "-"}
    if not term:
        raise ValueError(f"products 条目缺少 name 也缺少 unified_id：{entry!r}")

    res = st_get(token, "/v1/unified/search_entities",
                 {"entity_type": "app", "term": term, "limit": 5})
    apps = [a for a in res if a.get("is_unified") or a.get("os") == "unified"] or res
    if not apps:
        raise SystemExit(f"搜索『{term}』无结果，请换关键词或直接提供 unified_id。")
    top = apps[0]
    return {
        "unified_id": top.get("app_id") or top.get("id"),
        "name": top.get("name", term),
        "publisher": top.get("publisher_name", "?"),
        "release_date": (top.get("release_date") or "-")[:10],
        "candidates": [a.get("name") for a in apps[:5]],
    }


# ----------------------- 取数 -----------------------

def _split_countries(countries: List[str]) -> Tuple[List[str], bool]:
    reals = [c for c in countries if c.upper() != OTHER_LABEL]
    need_other = any(c.upper() == OTHER_LABEL for c in countries)
    return reals, need_other


def fetch_sales(token, app_ids, countries, start, end, gran) -> List[dict]:
    out: List[dict] = []
    for i in range(0, len(app_ids), 25):
        out.extend(st_get(token, "/v1/unified/sales_report_estimates", {
            "app_ids": ",".join(app_ids[i:i + 25]),
            "countries": ",".join(countries),
            "date_granularity": gran, "start_date": start, "end_date": end,
        }))
    return out


def fetch_dau(token, app_ids, countries, start, end) -> List[dict]:
    out: List[dict] = []
    for i in range(0, len(app_ids), 25):
        out.extend(st_get(token, "/v1/unified/usage/active_users", {
            "app_ids": ",".join(app_ids[i:i + 25]),
            "countries": ",".join(countries),
            "time_period": "day", "start_date": start, "end_date": end,
        }))
    return out


# ----------------------- 组装 DataFrame -----------------------

def _period_floor(ts: pd.Timestamp, gran: str) -> pd.Timestamp:
    if ts.tzinfo is not None:
        ts = ts.tz_localize(None)
    if gran == "monthly":
        return ts.to_period("M").to_timestamp()
    if gran == "quarterly":
        return ts.to_period("Q").to_timestamp()
    if gran == "weekly":
        return ts.to_period("W").to_timestamp()
    return ts.normalize()


def build_sales_df(sales, id2name, os_mode, gran) -> pd.DataFrame:
    dl_fields, rev_fields = OS_SALES_FIELDS[os_mode]
    rows = []
    for rec in sales:
        aid = str(rec.get("app_id"))
        rows.append({
            "Unified Name": id2name.get(aid, aid),
            "Country / Region": rec.get("country"),
            "Date": _period_floor(pd.to_datetime(rec.get("date")), gran),
            "Downloads": sum(rec.get(f, 0) or 0 for f in dl_fields),
            "Revenue ($)": sum(rec.get(f, 0) or 0 for f in rev_fields) / 100.0,  # 美分→美元
        })
    return pd.DataFrame(rows)


def derive_other(df: pd.DataFrame, real_countries: List[str]) -> pd.DataFrame:
    ww = df[df["Country / Region"] == WW_CODE]
    if ww.empty:
        print("   [warn] 需要 OTHER 但 WW 数据为空，跳过 OTHER 推导")
        return df[df["Country / Region"] != WW_CODE]
    keys = ["Unified Name", "Date"]
    reals = df[df["Country / Region"].isin(real_countries)]
    real_sum = reals.groupby(keys, as_index=False)[["Downloads", "Revenue ($)"]].sum()
    ww_sum = ww.groupby(keys, as_index=False)[["Downloads", "Revenue ($)"]].sum()
    m = ww_sum.merge(real_sum, on=keys, how="left", suffixes=("_ww", "_real")).fillna(0)
    m["Downloads"] = (m["Downloads_ww"] - m["Downloads_real"]).clip(lower=0)
    m["Revenue ($)"] = (m["Revenue ($)_ww"] - m["Revenue ($)_real"]).clip(lower=0)
    m["Country / Region"] = "Other"
    other = m[keys + ["Country / Region", "Downloads", "Revenue ($)"]]
    return pd.concat([df[df["Country / Region"] != WW_CODE], other], ignore_index=True)


def add_dau(df, dau_recs, id2name, os_mode, gran) -> pd.DataFrame:
    fields = OS_DAU_FIELDS[os_mode]
    rows = []
    for rec in dau_recs:
        aid = str(rec.get("app_id"))
        rows.append({
            "Unified Name": id2name.get(aid, aid),
            "Country / Region": rec.get("country"),
            "Date": _period_floor(pd.to_datetime(rec.get("date")), gran),
            "DAU": sum(rec.get(f, 0) or 0 for f in fields),
        })
    if not rows:
        return df
    dau_df = (pd.DataFrame(rows)
              .groupby(["Unified Name", "Country / Region", "Date"], as_index=False)["DAU"].mean())
    return df.merge(dau_df, on=["Unified Name", "Country / Region", "Date"], how="left")


# ----------------------- 口径调整 / 汇总 / Excel（自包含）-----------------------

def adjust(df: pd.DataFrame, caliber: str = "net") -> Tuple[pd.DataFrame, dict]:
    country_col, downloads_col, revenue_col = "Country / Region", "Downloads", "Revenue ($)"
    name_col = "Unified Name"
    dau_col = "DAU" if "DAU" in df.columns else None

    # 口径: net=不还原(share=1.0, ST 原始净收入); gross=÷0.7 还原平台分成(玩家毛支出)
    share = GROSS_SHARE if caliber == "gross" else 1.0
    is_cn = df[country_col].isin(CN_CODES)
    df["Is_CN"] = np.where(is_cn, "Y", "N")
    df["Downloads_adj"] = np.where(is_cn, df[downloads_col] * CN_DL_MULT, df[downloads_col])
    df["Revenue_adj ($)"] = np.where(
        is_cn, df[revenue_col] / share * CN_REV_MULT, df[revenue_col] / share)
    with np.errstate(divide="ignore", invalid="ignore"):
        df["RPD_adj ($)"] = np.where(
            df["Downloads_adj"] > 0, df["Revenue_adj ($)"] / df["Downloads_adj"], np.nan)

    has_dau = dau_col is not None
    if has_dau:
        df["DAU_adj"] = np.where(is_cn, df[dau_col] * CN_DAU_MULT, df[dau_col])
        with np.errstate(divide="ignore", invalid="ignore"):
            df["ARPDAU_adj ($)"] = np.where(
                df["DAU_adj"] > 0, df["Revenue_adj ($)"] / df["DAU_adj"], np.nan)

    return df, {"country_col": country_col, "name_col": name_col,
                "downloads_col": downloads_col, "revenue_col": revenue_col,
                "has_dau": has_dau, "is_cn_mask": is_cn}


def summarize_by(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    agg = (df.groupby(group_col, dropna=False)
           .agg(Downloads_adj=("Downloads_adj", "sum"),
                **{"Revenue_adj ($)": ("Revenue_adj ($)", "sum")},
                Rows=("Downloads_adj", "size")).reset_index())
    agg["RPD_blended ($)"] = np.where(
        agg["Downloads_adj"] > 0, agg["Revenue_adj ($)"] / agg["Downloads_adj"], np.nan)
    return agg.sort_values("Revenue_adj ($)", ascending=False).reset_index(drop=True)


def write_excel(detail, by_app, by_region, dst: Path) -> None:
    int_fmt, money_fmt, rate_fmt = "#,##0", "#,##0.00", "#,##0.0000"
    fmts = {
        "明细": {"Downloads": int_fmt, "Downloads_adj": int_fmt, "Revenue ($)": money_fmt,
               "Revenue_adj ($)": money_fmt, "RPD_adj ($)": rate_fmt, "DAU": int_fmt,
               "DAU_adj": int_fmt, "ARPDAU_adj ($)": rate_fmt},
        "按App汇总": {"Downloads_adj": int_fmt, "Revenue_adj ($)": money_fmt,
                   "RPD_blended ($)": rate_fmt, "Rows": int_fmt},
        "按地区汇总": {"Downloads_adj": int_fmt, "Revenue_adj ($)": money_fmt,
                   "RPD_blended ($)": rate_fmt, "Rows": int_fmt},
    }
    with pd.ExcelWriter(dst, engine="openpyxl") as writer:
        detail.to_excel(writer, sheet_name="明细", index=False)
        by_app.to_excel(writer, sheet_name="按App汇总", index=False)
        by_region.to_excel(writer, sheet_name="按地区汇总", index=False)
        wb = writer.book
        for sheet, col_fmts in fmts.items():
            ws = wb[sheet]
            header = [c.value for c in ws[1]]
            for col_name, fmt in col_fmts.items():
                if col_name in header:
                    letter = get_column_letter(header.index(col_name) + 1)
                    for cell in ws[letter][1:]:
                        cell.number_format = fmt
            ws.freeze_panes = "A2"


# ----------------------- 主流程 -----------------------

def load_query(path: Path, cfg: dict) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        q = json.load(f)
    for k in ("products", "start_date", "end_date"):
        if not q.get(k):
            raise SystemExit(f"query.json 缺少必填字段：{k}")
    q.setdefault("countries", cfg.get("default_countries", ["CN", "US", "JP", "KR", "OTHER"]))
    q.setdefault("granularity", cfg.get("default_granularity", "monthly"))
    q.setdefault("os", cfg.get("default_os", "unified"))
    q.setdefault("include_dau", cfg.get("include_dau_default", False))
    q.setdefault("caliber", cfg.get("default_caliber", "net"))
    if q["caliber"] not in ("net", "gross"):
        raise SystemExit(f"caliber 必须是 net 或 gross，收到：{q['caliber']}")
    if q["os"] not in OS_SALES_FIELDS:
        raise SystemExit(f"os 必须是 {list(OS_SALES_FIELDS)} 之一，收到：{q['os']}")
    return q


def main(argv: list) -> None:
    args = [a for a in argv[1:] if not a.startswith("--")]
    flags = {a for a in argv[1:] if a.startswith("--")}
    if not args:
        print(__doc__)
        sys.exit(1)

    query_path = Path(args[0]).expanduser().resolve()
    if not query_path.exists():
        raise FileNotFoundError(query_path)

    cfg = load_config()
    token = get_token(cfg)
    q = load_query(query_path, cfg)

    print(f"[resolve] 解析 {len(q['products'])} 个产品 …")
    resolved = [resolve_product(token, p) for p in q["products"]]
    id2name: Dict[str, str] = {}
    for r in resolved:
        id2name[str(r["unified_id"])] = r["name"]
        cand = f"  候选: {r['candidates']}" if r.get("candidates") else ""
        print(f"   ✓ {r['name']}  (发行商={r['publisher']}, 上线={r['release_date']}, "
              f"unified_id={r['unified_id']}){cand}")

    if "--resolve-only" in flags:
        print("\n[resolve-only] 仅解析，未拉数。确认匹配无误后去掉 --resolve-only 重跑。")
        return

    os_mode, gran = q["os"], q["granularity"]
    real_countries, need_other = _split_countries(q["countries"])
    fetch_countries = list(real_countries)
    if need_other and WW_CODE not in fetch_countries:
        fetch_countries.append(WW_CODE)

    app_ids = [str(r["unified_id"]) for r in resolved]
    print(f"\n[fetch] sales: os={os_mode} gran={gran} 国家={fetch_countries} "
          f"区间={q['start_date']}~{q['end_date']} 口径={q['caliber']}")
    sales = fetch_sales(token, app_ids, fetch_countries, q["start_date"], q["end_date"], gran)
    print(f"   ← {len(sales):,} 条 sales 记录")
    if not sales:
        raise SystemExit("API 返回 0 条 sales。检查产品/国家/日期，或该产品在所选区间无数据。")

    df = build_sales_df(sales, id2name, os_mode, gran)
    if need_other:
        df = derive_other(df, real_countries)
        print(f"   [other] 已按 WW − {real_countries} 推导『Other』")

    if q["include_dau"]:
        print("[fetch] DAU: unified/usage/active_users (day 粒度 → period 均值)")
        dau_recs = fetch_dau(token, app_ids, fetch_countries, q["start_date"], q["end_date"])
        print(f"   ← {len(dau_recs):,} 条 DAU 记录")
        df = add_dau(df, dau_recs, id2name, os_mode, gran)
        if need_other:
            print("   [note] DAU 不做 Other 推导，Other 行 DAU 留空")

    df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%Y-%m-%d")

    df, meta = adjust(df, q["caliber"])
    by_app = summarize_by(df, meta["name_col"])
    by_region = summarize_by(df, meta["country_col"])

    dst = (Path(args[1]).expanduser().resolve() if len(args) >= 2
           else query_path.with_name(query_path.stem + "_ST_adjusted.xlsx"))
    write_excel(df, by_app, by_region, dst)

    print()
    print("=" * 60)
    print(f"[done] 写入: {dst}")
    print(f"[qa] 明细行数: {len(df):,}   产品数: {df['Unified Name'].nunique()}   "
          f"DAU/ARPDAU: {'已生成' if meta['has_dau'] else '未生成'}")
    print(f"[qa] CN 行数: {int(meta['is_cn_mask'].sum())}")
    print("\n[qa] 按地区:")
    print(by_region.to_string(index=False))
    print("\n[qa] 按 App (按 Revenue_adj):")
    print(by_app.to_string(index=False))
    caliber_desc = "原始净收入(不÷0.7)" if q["caliber"] == "net" else "÷0.7 还原成毛支出(gross)"
    print(f"\n[hint] 口径={q['caliber']}: _adj 列 = {caliber_desc}。CN上修默认关。切换改 query.json 的 caliber 字段(net/gross)。")


if __name__ == "__main__":
    main(sys.argv)
